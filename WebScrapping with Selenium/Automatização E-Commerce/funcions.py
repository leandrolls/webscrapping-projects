import openpyxl
import win32com.client as win32


def criarPlanilha(nomes, precos):
    index = 2
    planilha = openpyxl.Workbook()
    produto = planilha['Sheet']

    produto.title = 'Produto'
    produto['A1'] = 'Nome'
    produto['B1'] = 'Preço'

    for nome, preco in zip(nomes, precos):
        produto.cell(column=1, row=index, value=nome)
        produto.cell(column=2, row=index, value=preco)
        index = index + 1
        planilha.save("planilha_de_preços_e-commerce.xlsx")
    print("Concluido!")


def enviar_email():
    outlook = win32.Dispatch('outlook.application')
    email = outlook.CreateItem(0)

    email.To = "joe.limals.14@gmail.com; leandrolssp.eng@gmail.com"
    email.Subject = "Relatório de preços"
    email.HTMLBody = f""""
    <p> Bom dia! 
    Segue em anexo planilha com preços de produtos nos e-commercers parceiros! </p>
    """
    anexo = "E:\Documentos\Carreira\Python Projetos\Automatização E-Commerce\planilha_de_preços_e-commerce.xlsx"
    email.Attachments.Add(anexo)

    email.Send()
    print("E-mail enviado")
